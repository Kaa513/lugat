"""Import HSK words from Excel into dictionary.db"""

import sys
from openpyxl import load_workbook
from pathlib import Path

# Путь к БД
sys.path.insert(0, str(Path(__file__).parent))
from database import get_connection

EXCEL_PATH = Path(__file__).parent / "hsk_old_levels.xlsx"

HSK_LEVEL_MAP = {
    "hsk_1": 1, "hsk_2": 2, "hsk_3": 3,
    "hsk_4": 4, "hsk_5": 5, "hsk_6": 6,
}

def import_hsk():
    wb = load_workbook(EXCEL_PATH, read_only=True)
    total_added = 0
    total_updated = 0
    total_skipped = 0

    with get_connection() as conn:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            print(f"\n[{sheet_name}] Колонки: {headers}")

            added = updated = skipped = 0

            for row in rows[1:]:
                if not any(row):
                    continue

                data = dict(zip(headers, row))

                chinese = (data.get("chinese") or "").strip()
                if not chinese:
                    skipped += 1
                    continue

                pinyin = (data.get("pinyin") or "").strip()
                english = (data.get("english") or data.get("english ") or "").strip()
                uzbek = (data.get("uzbek") or "").strip()
                hsk_raw = (str(data.get("hsk_level") or "")).strip().lower()
                hsk_level = HSK_LEVEL_MAP.get(hsk_raw)

                # Объединяем словосочетание и предложение через " | "
                ex_cn_parts = []
                ex_uz_parts = []
                ex_en_parts = []

                if data.get("example_chinese"):
                    ex_cn_parts.append(str(data["example_chinese"]).strip())
                if data.get("sentence_example_chinese"):
                    ex_cn_parts.append(str(data["sentence_example_chinese"]).strip())

                if data.get("example_uzbek"):
                    ex_uz_parts.append(str(data["example_uzbek"]).strip())
                if data.get("sentence_example_uzbek"):
                    ex_uz_parts.append(str(data["sentence_example_uzbek"]).strip())

                if data.get("example_english"):
                    ex_en_parts.append(str(data["example_english"]).strip())
                if data.get("sentence_example_english"):
                    ex_en_parts.append(str(data["sentence_example_english"]).strip())

                example_chinese = " | ".join(ex_cn_parts) if ex_cn_parts else None
                example_uzbek = " | ".join(ex_uz_parts) if ex_uz_parts else None

                # Проверяем есть ли уже такое слово
                existing = conn.execute(
                    "SELECT id FROM words WHERE chinese = ?", (chinese,)
                ).fetchone()

                if existing:
                    conn.execute(
                        """UPDATE words SET pinyin=?, english=?, uzbek=?,
                           hsk_level=?, example_chinese=?, example_uzbek=?
                           WHERE chinese=?""",
                        (pinyin, english, uzbek, hsk_level,
                         example_chinese, example_uzbek, chinese)
                    )
                    updated += 1
                else:
                    conn.execute(
                        """INSERT INTO words
                           (chinese, pinyin, english, uzbek, hsk_level,
                            example_chinese, example_uzbek)
                           VALUES (?,?,?,?,?,?,?)""",
                        (chinese, pinyin, english, uzbek, hsk_level,
                         example_chinese, example_uzbek)
                    )
                    added += 1

            conn.commit()
            print(f"  Добавлено: {added}, Обновлено: {updated}, Пропущено: {skipped}")
            total_added += added
            total_updated += updated
            total_skipped += skipped

    print(f"\nИТОГО: добавлено {total_added}, обновлено {total_updated}, пропущено {total_skipped}")

if __name__ == "__main__":
    import_hsk()
